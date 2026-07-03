"""XLSX adapter (openpyxl).

Spreadsheets are already a grid, so the work is mostly faithful transcription:
one sheet -> one page, merged header cells propagated across their span so
multi-column headers survive, values coerced to strings.
"""
from __future__ import annotations

from typing import List, Optional

from openpyxl import load_workbook

from ..model.elements import Page
from ..model.table import ExtractedTable
from .base import DocumentAdapter, register

_MAX_ROWS = 5000
_MAX_COLS = 200


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


@register
class XlsxAdapter(DocumentAdapter):
    extensions = ("xlsx", "xlsm")

    def parse(self, path: str) -> List[Page]:
        wb = load_workbook(path, data_only=True)
        pages: List[Page] = []
        for si, ws in enumerate(wb.worksheets):
            max_row = min(ws.max_row or 0, _MAX_ROWS)
            max_col = min(ws.max_column or 0, _MAX_COLS)
            if max_row == 0 or max_col == 0:
                pages.append(Page(index=si, width=0, height=0, source_kind="xlsx"))
                continue
            grid = [[_s(ws.cell(r, c).value) for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]
            # propagate merged header/label cells across their range
            for rng in ws.merged_cells.ranges:
                tl = _s(ws.cell(rng.min_row, rng.min_col).value)
                for r in range(rng.min_row, min(rng.max_row, max_row) + 1):
                    for c in range(rng.min_col, min(rng.max_col, max_col) + 1):
                        grid[r - 1][c - 1] = tl
            grid = _trim(grid)
            table = _grid_to_table(grid, si, title=ws.title)
            pages.append(
                Page(index=si, width=0, height=0, source_kind="xlsx", native_tables=[table] if table else [])
            )
        return pages


def _trim(grid: List[List[Optional[str]]]) -> List[List[Optional[str]]]:
    while grid and all(v is None for v in grid[-1]):
        grid.pop()
    while grid and all(v is None for v in grid[0]):
        grid.pop(0)
    if not grid:
        return grid
    ncol = len(grid[0])
    last = 0
    for row in grid:
        for c in range(ncol - 1, -1, -1):
            if c < len(row) and row[c] is not None:
                last = max(last, c)
                break
    return [row[: last + 1] for row in grid]


def _grid_to_table(grid: List[List[Optional[str]]], page_index: int, title: Optional[str]) -> Optional[ExtractedTable]:
    if not grid:
        return None
    header = grid[0]
    body = grid[1:]
    return ExtractedTable(
        title=title,
        row_headers=[""] * len(body),
        col_headers=[h or f"col{i + 1}" for i, h in enumerate(header)],
        cells=[list(r) for r in body],
        source_page=page_index,
        kind="grid",
    )
